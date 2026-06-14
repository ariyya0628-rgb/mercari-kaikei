import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from src.mercari_converter import convert_sold_csv_to_excel


CSV_TEXT = """購入完了日,出品日時,購入日時,更新日時,取引情報URL,商品ID,商品名,商品URL,商品代金,販売手数料,配送料,他費用,税率,販売利益,配送料の負担,配送の方法,発送元の地域,発送までの日数,商品の状態,いいね！
2026/06/03 10:00,2026/06/01 09:00,2026/06/02 12:00,2026/06/03 10:00,https://example.com/tx/1,m111,商品A,https://example.com/item/1,"1,200",120,210,0,10%,870,送料込み,らくらくメルカリ便,東京都,1〜2日で発送,目立った傷や汚れなし,3
,2026/06/04 09:00,2026/06/05 12:30,2026/06/06 13:00,https://example.com/tx/2,m222,商品B,https://example.com/item/2,800,80,0,10,10%,710,着払い,ゆうゆうメルカリ便,大阪府,2〜3日で発送,新品、未使用,1
"""


class MercariConverterTest(unittest.TestCase):
    def test_convert_sold_csv_creates_accounting_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            input_csv = tmp_path / "mercari_sold_sample.csv"
            output_xlsx = tmp_path / "mercari_accounting.xlsx"
            input_csv.write_text(CSV_TEXT, encoding="utf-8-sig")

            convert_sold_csv_to_excel(input_csv, output_xlsx)

            workbook = load_workbook(output_xlsx, data_only=False)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "README",
                    "sold_raw",
                    "売上ベース",
                    "売上金確定ベース",
                    "残高消込表",
                    "メルカード確認",
                    "銀行振込確認",
                    "月次集計",
                ],
            )

            sales = workbook["売上ベース"]
            self.assertEqual(sales["A2"].value, sales["B2"].value)
            self.assertEqual(sales["I2"].value, 1200)
            self.assertEqual(sales.freeze_panes, "A2")

            settled = workbook["売上金確定ベース"]
            self.assertIsNone(settled["B3"].value)
            self.assertEqual(settled["A3"].value, settled["C3"].value)

            balance = workbook["残高消込表"]
            self.assertEqual(balance["B2"].value, "売上金確定")
            self.assertEqual(balance["D2"].value, 870)
            self.assertEqual(balance["F2"].value, "=D2-E2")
            self.assertEqual(balance["F3"].value, "=F2+D3-E3")

            mercard = workbook["メルカード確認"]
            self.assertEqual(mercard["D2"].value, "=B2-C2")
            self.assertEqual(mercard["E2"].value, "=B2-C2-D2")

            monthly = workbook["月次集計"]
            rows = list(monthly.iter_rows(values_only=True))
            self.assertIn(("売上ベース", "2026-06", 2, 2000, 200, 210, 10, 1580), rows)
            self.assertIn(("売上金確定ベース", "2026-06", 2, 2000, 200, 210, 10, 1580), rows)


if __name__ == "__main__":
    unittest.main()
