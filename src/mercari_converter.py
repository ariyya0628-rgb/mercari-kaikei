from __future__ import annotations

import csv
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


SHEET_NAMES = [
    "README",
    "sold_raw",
    "売上ベース",
    "売上金確定ベース",
    "残高消込表",
    "メルカード確認",
    "銀行振込確認",
    "月次集計",
]

DATE_COLUMNS = {"購入完了日", "出品日時", "購入日時", "更新日時"}
AMOUNT_COLUMNS = {"商品代金", "販売手数料", "配送料", "他費用", "販売利益"}
INPUT_ENCODINGS = ("utf-8-sig", "utf-8", "cp932", "shift_jis")


def convert_sold_csv_to_excel(input_csv: str | Path, output_xlsx: str | Path) -> Path:
    input_path = Path(input_csv)
    output_path = Path(output_xlsx)
    raw_headers, raw_rows = read_sold_csv(input_path)
    rows = [normalize_row(row) for row in raw_rows]

    workbook = Workbook()
    workbook.remove(workbook.active)

    build_readme_sheet(workbook)
    build_raw_sheet(workbook, raw_headers, raw_rows)
    build_sales_sheet(workbook, rows)
    settled_rows = build_settled_sheet(workbook, rows)
    build_balance_sheet(workbook, settled_rows)
    build_mercard_sheet(workbook)
    build_bank_sheet(workbook)
    build_monthly_sheet(workbook, rows, settled_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def read_sold_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    last_error: Exception | None = None
    for encoding in INPUT_ENCODINGS:
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                reader = csv.DictReader(file)
                headers = list(reader.fieldnames or [])
                rows = [{key: (value or "") for key, value in row.items()} for row in reader]
            if not headers:
                raise ValueError("CSVのヘッダー行を読み取れませんでした。")
            return headers, rows
        except UnicodeDecodeError as error:
            last_error = error
    raise UnicodeError(f"CSVの文字コードを判定できませんでした: {path}") from last_error


def normalize_row(row: dict[str, str]) -> dict[str, Any]:
    normalized: dict[str, Any] = dict(row)
    for column in DATE_COLUMNS:
        normalized[column] = parse_date(row.get(column, ""))
    for column in AMOUNT_COLUMNS:
        normalized[column] = parse_amount(row.get(column, ""))
    return normalized


def parse_date(value: Any) -> datetime | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("-", "/")
    formats = (
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d",
        "%Y年%m月%d日 %H:%M",
        "%Y年%m月%d日",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def parse_amount(value: Any) -> int | float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    cleaned = (
        text.replace(",", "")
        .replace("¥", "")
        .replace("￥", "")
        .replace("円", "")
        .strip()
    )
    if not cleaned:
        return None
    try:
        number = float(cleaned)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def build_readme_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("README")
    sheet["A1"] = "メルカリ売上 会計資料"
    sheet["A1"].font = Font(size=16, bold=True)
    lines = [
        "このExcelは、メルカリのsold CSVをもとに、売上発生、売上金確定、メルカリ残高の増減を税理士へ説明するための資料です。",
        "",
        "メルカリでは、購入日時と売上金確定日は一致しない場合があります。",
        "購入者の受取評価後に、販売手数料や送料を差し引いた販売利益が確定します。",
        "確定した販売利益はメルカリ残高に加算されます。",
        "メルカリ残高は、銀行振込だけでなく、メルカード支払い・メルペイ利用などに使われる場合があります。",
        "そのため、この資料では売上ベース、売上金確定ベース、残高消込表を分けて管理します。",
        "",
        "各シートの見方",
        "sold_raw: メルカリからダウンロードしたsold CSVの内容をできるだけそのまま保持した確認用シートです。",
        "売上ベース: 商品が売れた日である購入日時を基準に並べた資料です。",
        "売上金確定ベース: 購入完了日を基準に、販売利益がメルカリ残高へ加算された日を示す資料です。購入完了日が空欄の場合は更新日時を使用します。",
        "残高消込表: sold CSV由来の販売利益を増加として自動反映し、銀行振込・メルカード充当・メルペイ利用などの減少を手入力で追記するための表です。",
        "メルカード確認: スクリーンショット明細を見ながら、請求額・銀行引落額・残高充当額の関係を手入力で確認するシートです。",
        "銀行振込確認: 銀行口座へ実際に入金された金額と、残高消込表の銀行振込行を照合するためのシートです。",
        "月次集計: 売上ベースと売上金確定ベースを月別に集計したシートです。",
    ]
    for row_number, line in enumerate(lines, start=3):
        sheet.cell(row=row_number, column=1, value=line)
    sheet.column_dimensions["A"].width = 120
    sheet.freeze_panes = "A3"
    sheet["A3"].alignment = Alignment(wrap_text=True, vertical="top")


def build_raw_sheet(workbook: Workbook, headers: list[str], rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet("sold_raw")
    write_table(sheet, headers, [[row.get(header, "") for header in headers] for row in rows])
    apply_sheet_format(sheet)


def build_sales_sheet(workbook: Workbook, rows: list[dict[str, Any]]) -> None:
    columns = ["基準日", "購入日時", "購入完了日", "更新日時", "商品ID", "商品名", "商品URL", "取引情報URL", "商品代金", "販売手数料", "配送料", "他費用", "販売利益", "配送の方法", "商品の状態"]
    sorted_rows = sorted(rows, key=lambda row: sort_key(row.get("購入日時")))
    table_rows = []
    for row in sorted_rows:
        table_rows.append([row.get("購入日時"), row.get("購入日時"), row.get("購入完了日"), row.get("更新日時"), row.get("商品ID"), row.get("商品名"), row.get("商品URL"), row.get("取引情報URL"), row.get("商品代金"), row.get("販売手数料"), row.get("配送料"), row.get("他費用"), row.get("販売利益"), row.get("配送の方法"), row.get("商品の状態")])
    sheet = workbook.create_sheet("売上ベース")
    write_table(sheet, columns, table_rows)
    apply_sheet_format(sheet, amount_columns={9, 10, 11, 12, 13}, date_columns={1, 2, 3, 4})


def build_settled_sheet(workbook: Workbook, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    columns = ["基準日", "購入完了日", "更新日時", "購入日時", "商品ID", "商品名", "商品URL", "取引情報URL", "商品代金", "販売手数料", "配送料", "他費用", "販売利益", "配送の方法", "商品の状態"]
    enriched_rows = []
    for row in rows:
        enriched = dict(row)
        enriched["基準日"] = row.get("購入完了日") or row.get("更新日時")
        enriched_rows.append(enriched)
    sorted_rows = sorted(enriched_rows, key=lambda row: sort_key(row.get("基準日")))
    table_rows = []
    for row in sorted_rows:
        table_rows.append([row.get("基準日"), row.get("購入完了日"), row.get("更新日時"), row.get("購入日時"), row.get("商品ID"), row.get("商品名"), row.get("商品URL"), row.get("取引情報URL"), row.get("商品代金"), row.get("販売手数料"), row.get("配送料"), row.get("他費用"), row.get("販売利益"), row.get("配送の方法"), row.get("商品の状態")])
    sheet = workbook.create_sheet("売上金確定ベース")
    write_table(sheet, columns, table_rows)
    apply_sheet_format(sheet, amount_columns={9, 10, 11, 12, 13}, date_columns={1, 2, 3, 4})
    return sorted_rows


def build_balance_sheet(workbook: Workbook, rows: list[dict[str, Any]]) -> None:
    columns = ["日付", "区分", "内容", "増加", "減少", "残高", "商品ID", "取引情報URL", "証憑", "備考"]
    table_rows = []
    for row in rows:
        table_rows.append([row.get("基準日"), "売上金確定", row.get("商品名"), row.get("販売利益") or 0, 0, None, row.get("商品ID"), row.get("取引情報URL"), row.get("取引情報URL"), "sold CSVより自動反映"])
    for _ in range(50):
        table_rows.append([None] * 10)
    sheet = workbook.create_sheet("残高消込表")
    write_table(sheet, columns, table_rows)
    for row_number in range(2, sheet.max_row + 1):
        sheet.cell(row=row_number, column=6, value=f"=D{row_number}-E{row_number}" if row_number == 2 else f"=F{row_number - 1}+D{row_number}-E{row_number}")
    validation = DataValidation(type="list", formula1='"売上金確定,銀行振込,メルカード充当,メルペイ利用,その他調整"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"B2:B{sheet.max_row}")
    apply_sheet_format(sheet, amount_columns={4, 5, 6}, date_columns={1})


def build_mercard_sheet(workbook: Workbook) -> None:
    columns = ["対象月", "メルカード請求金額", "銀行口座引落額", "メルカリ残高充当額", "差額確認", "支払日", "スクショファイル名", "備考"]
    rows = [[None] * len(columns) for _ in range(24)]
    sheet = workbook.create_sheet("メルカード確認")
    write_table(sheet, columns, rows)
    for row_number in range(2, sheet.max_row + 1):
        sheet.cell(row=row_number, column=4, value=f"=B{row_number}-C{row_number}")
        sheet.cell(row=row_number, column=5, value=f"=B{row_number}-C{row_number}-D{row_number}")
    apply_sheet_format(sheet, amount_columns={2, 3, 4, 5}, date_columns={6})


def build_bank_sheet(workbook: Workbook) -> None:
    columns = ["振込日", "入金額", "銀行名", "口座", "摘要", "対応する残高消込表の行", "証憑", "備考"]
    rows = [[None] * len(columns) for _ in range(50)]
    sheet = workbook.create_sheet("銀行振込確認")
    write_table(sheet, columns, rows)
    apply_sheet_format(sheet, amount_columns={2}, date_columns={1})


def build_monthly_sheet(workbook: Workbook, sales_rows: list[dict[str, Any]], settled_rows: list[dict[str, Any]]) -> None:
    columns = ["区分", "月", "件数", "商品代金合計", "販売手数料合計", "配送料合計", "他費用合計", "販売利益合計"]
    output_rows = []
    output_rows.extend(aggregate_monthly("売上ベース", sales_rows, "購入日時"))
    output_rows.extend(aggregate_monthly("売上金確定ベース", settled_rows, "基準日"))
    sheet = workbook.create_sheet("月次集計")
    write_table(sheet, columns, output_rows)
    apply_sheet_format(sheet, amount_columns={4, 5, 6, 7, 8})


def aggregate_monthly(label: str, rows: list[dict[str, Any]], date_column: str) -> list[list[Any]]:
    buckets: dict[str, dict[str, Any]] = defaultdict(lambda: {"件数": 0, "商品代金": 0, "販売手数料": 0, "配送料": 0, "他費用": 0, "販売利益": 0})
    for row in rows:
        date_value = row.get(date_column)
        if not isinstance(date_value, datetime):
            continue
        month = date_value.strftime("%Y-%m")
        bucket = buckets[month]
        bucket["件数"] += 1
        for column in AMOUNT_COLUMNS:
            bucket[column] += row.get(column) or 0
    return [[label, month, bucket["件数"], bucket["商品代金"], bucket["販売手数料"], bucket["配送料"], bucket["他費用"], bucket["販売利益"]] for month, bucket in sorted(buckets.items())]


def write_table(sheet: Any, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)


def apply_sheet_format(sheet: Any, amount_columns: set[int] | None = None, date_columns: set[int] | None = None) -> None:
    amount_columns = amount_columns or set()
    date_columns = date_columns or set()
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    if sheet.max_column:
        sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        column_index = column_cells[0].column
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        width = min(max((len(value) for value in values), default=8) + 2, 60)
        sheet.column_dimensions[column_cells[0].column_letter].width = width
        for cell in column_cells[1:]:
            if column_index in amount_columns:
                cell.number_format = '#,##0'
            if column_index in date_columns:
                cell.number_format = "yyyy/mm/dd hh:mm"
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"


def sort_key(value: Any) -> tuple[int, datetime]:
    if isinstance(value, datetime):
        return (0, value)
    return (1, datetime.max)
